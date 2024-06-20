import wx
from sympy import Matrix
import numpy as np
from openpyxl import Workbook, load_workbook
import os


#-------------------------UI部分-------------------------------

class ArgumentSelect(wx.Frame):
    def __init__(self):
        super().__init__(None, title="请选择你的培养基", size=(300, 320))
        
        # 创建菜单栏
        menu_bar = wx.MenuBar()
        file_menu = wx.Menu()
        open_item = file_menu.Append(wx.ID_OPEN, "说明", "Open a file")
        exit_item = file_menu.Append(wx.ID_EXIT, "Exit", "Exit the application")
        menu_bar.Append(file_menu, "帮助")      # 将菜单添加到菜单栏
        self.SetMenuBar(menu_bar)               # 将菜单栏设置为窗口的菜单栏
        self.Bind(wx.EVT_MENU, self.on_open, open_item)
        self.Bind(wx.EVT_MENU, self.on_exit, exit_item)
        

        self.panel = wx.Panel(self)
        self.sizer = wx.GridSizer(rows=7, cols=5, hgap=5, vgap=5)
        
        # self.param1_label = wx.StaticText(self.panel, label="矩阵维数：")
        # self.cube_num = wx.Choice(self.panel, choices=[str(i) for i in range(2, 9)])
        
        # self.param2_label = wx.StaticText(self.panel, label="矩阵个数:")
        self.sheetlist = []
        for sheet_row in range(6):
            for sheet_col in range(5):
                sheet_name = chr(sheet_col + 65) + str(sheet_row + 1)
                self.sheetlist.append(sheet_name)
        # self.state_num = wx.Choice(self.panel, choices=[str(i) for i in self.sheetlist])
        
        # self.shut_button = wx.Button(self.panel, label="关闭其他窗口")
        # self.shut_button.Bind(wx.EVT_BUTTON, self.on_close_button)
        
        for sheet_name in self.sheetlist:
            button = wx.Button(self.panel, label=sheet_name)
            button.Bind(wx.EVT_BUTTON, self.on_button_click)
            self.sizer.Add(button, 0, wx.EXPAND)

        sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.ok_button = wx.Button(self.panel, label="确定")
        self.ok_button.Bind(wx.EVT_BUTTON, self.on_ok)

        # self.sizer.Add(self.param1_label, 0, wx.EXPAND | wx.ALL, 5)
        # self.sizer.Add(self.cube_num, 0, wx.EXPAND | wx.ALL, 5)
        # self.sizer.Add(self.param2_label, 0, wx.EXPAND | wx.ALL, 5)
        # self.sizer.Add(self.state_num, 0, wx.EXPAND | wx.ALL, 5)
        self.sizer.Add(self.ok_button, 0, wx.ALIGN_CENTER | wx.ALL, 5)
        # self.sizer.Add(self.shut_button, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.panel.SetSizer(self.sizer)
        self.Show()

    def on_ok(self, event):
        # 读取 MatrixSum.xlsx 文件
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_file = os.path.join(self.current_dir, "MatrixSum.xlsx")
        wb = load_workbook(self.excel_file)

        # 创建一个空的三维 NumPy 数组
        matrix_3d = np.zeros((30, 6, 6))

        # 遍历每个 sheet,并将数据填充到三维矩阵中
        sheet_name_num = 0
        for sheet_name in self.sheetlist:
            sheet = wb[sheet_name]
            for row in range(6):
                for col in range(6):
                    matrix_3d[sheet_name_num, row, col] = sheet.cell(row=row+1, column=col+1).value
            sheet_name_num += 1

        result_matrix = np.zeros((6, 6))
        sheet_name_num = 0
        for row in range(6):
            for col in range(6):
                colony_sum = 0
                for sheet_name in self.sheetlist:
                    sheet = wb[sheet_name]
                    colony_sum += eval(sheet.cell(row=row+1, column=col+1).value)
                    sheet_name_num += 1
                result_matrix[row, col] = colony_sum
        print(result_matrix)

    def on_button_click(self, event):
        self.button = event.GetEventObject()
        
        self.Matrix_Frame = MatrixFrame(self.button.GetLabel())
        self.Matrix_Frame.Show()        
    
    def on_open(self, event):        # 处理打开菜单项的事件
        new_window = wx.Frame(None, title="Introduction", size=(425, 500), style=wx.DEFAULT_FRAME_STYLE ^ wx.RESIZE_BORDER)
        text = """		    【帮助】\n\n
我爱微生物学实验（真切）
                                    xy\n
                                2024年6月19日
        """
        text_ctrl = wx.TextCtrl(new_window, style=wx.TE_MULTILINE | wx.TE_READONLY)
        text_ctrl.SetValue(text)
        font = wx.Font(12, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, faceName="黑体")
        text_ctrl.SetFont(font)
        new_window.Show()
        
    def on_exit(self, event):        # 处理退出菜单项的事件
        self.Close()

    def on_close_button(self, event): # 关闭其他窗口
        top_windows = wx.GetTopLevelWindows()
        current_window = self
        for window in top_windows:
            if window != current_window:
                window.Close()



class MatrixFrame(wx.Frame):
    def __init__(self, sheet_name):
        # 表格
        self.sheet_name = sheet_name
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_file = os.path.join(self.current_dir, "MatrixSum.xlsx")
        # 检查是否存在 Excel 文件
        if os.path.exists(self.excel_file):
            self.workbook = load_workbook(self.excel_file)
        else:
            self.create_sheet()
        # 检查是否存在名为 sheet_name 的工作表
        if self.sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[self.sheet_name]
        else:
            self.worksheet = self.workbook.create_sheet(title=self.sheet_name, index=0)
        self.sheet_matrix = []
        for row in self.worksheet.iter_rows():
            row_data = [cell.value for cell in row]
            self.sheet_matrix.append(row_data)
        self.vary_matrix = self.sheet_matrix


        super().__init__(None, title=f"{sheet_name}", size=(400, 300))

        # 获取屏幕的尺寸
        screen_width, screen_height = wx.GetDisplaySize()
        # 计算窗口的左上角坐标，使其位于屏幕正中间
        window_width, window_height = self.GetSize()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        # 设置窗口的位置
        self.SetPosition((x, y))

        self.matrix_size = 6  # 矩阵大小
        self.matrix = [[False for _ in range(self.matrix_size)] for _ in range(self.matrix_size)]  # 创建一个默认大小的矩阵
        self.cell_size = 50  # 每个单元格的大小
        
        self.panel = wx.Panel(self)
        self.sizer = wx.GridSizer(self.matrix_size + 4, self.matrix_size + 1, 0, 0)
        
        # 添加列名称
        self.sizer.Add(wx.StaticText(self.panel), 0, wx.EXPAND)
        for i in range(self.matrix_size):
            col_label = i + 1
            self.sizer.Add(wx.StaticText(self.panel, label=str(col_label), style=wx.ALIGN_CENTER), 0, wx.EXPAND)
        
        # 添加行名称和单元格
        for i in range(self.matrix_size):
            row_label = chr(ord('A') + i)
            self.sizer.Add(wx.StaticText(self.panel, label=row_label, style=wx.ALIGN_CENTER), 0, wx.EXPAND)
            for j in range(self.matrix_size):
                toggle_btn = wx.ToggleButton(self.panel, id=wx.ID_ANY, label="", size=(self.cell_size, self.cell_size))
                toggle_btn.row = i  # 将行索引保存在toggle_btn的属性中
                toggle_btn.col = j  # 将列索引保存在toggle_btn的属性中
                # toggle_btn.colony_num = chr(i + 65) + str(j + 1)
                toggle_btn.Bind(wx.EVT_TOGGLEBUTTON, self.on_toggle)
                self.sizer.Add(toggle_btn, 0, wx.EXPAND)
                # 根据 sheet_matrix 中的值设置按钮的状态
                toggle_btn.SetValue(eval(self.sheet_matrix[i][j]) == 1)

        # 添加确定按钮
        self.confirm_button = wx.Button(self.panel, label="保存并退出")
        self.confirm_button.Bind(wx.EVT_BUTTON, self.on_ok)
        self.sizer.Add(self.confirm_button, 3, wx.EXPAND)

        self.panel.SetSizer(self.sizer)
        
    def create_sheet(self):
        # 获取当前文件所在的目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        # 创建新的工作簿
        self.workbook = Workbook()
        # 创建名为A1到E6的工作表
        for sheet_col in range(5):
            for sheet_row in range(6):
                sheet_name = chr(sheet_col + 65) + str(sheet_row + 1)
                worksheet = self.workbook.create_sheet(title=sheet_name, index=0)
                # 将每个单元格的值设置为 0
                for row in worksheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=6):
                    for cell in row:
                        cell.value = '0'
        # 保存Excel文件
        self.workbook.save(os.path.join(current_dir, "MatrixSum.xlsx"))

    def on_toggle(self, event): 
        toggle_btn = event.GetEventObject()
        row = toggle_btn.row
        col = toggle_btn.col
        self.matrix[row][col] = toggle_btn.GetValue()
        
        for i in range(self.matrix_size):
            for j in range(self.matrix_size):
                if self.matrix[i][j]:
                    self.vary_matrix[i][j] = 1

        self.panel.Layout()  # 更新面板布局，实现实时反馈

    def on_ok(self, event):
        self.vary_matrix = Matrix(self.vary_matrix)
        # 将 self.vary_matrix 中的数据写入工作表
        for row in range(self.matrix_size):
            for col in range(self.matrix_size):
                self.worksheet.cell(row=row+1, column=col+1, value=str(self.vary_matrix[row, col]))

        # 保存 Excel 文件
        self.workbook.save(self.excel_file)
        print("Excel file updated successfully!")

        top_windows = wx.GetTopLevelWindows()
        current_window = self
        for window in top_windows:
            if window == current_window:
                window.Close()


#------------------------计算部分----------------------------
            
def caculation():

    return 0

# --------------------------主体----------------------------
app = wx.App()
frame = ArgumentSelect()
frame.Show()
app.MainLoop()